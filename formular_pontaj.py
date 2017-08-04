import calendar
import datetime
import openpyxl
import os
import sys
import time
import timeit

from colorama import Fore
from colorama import init
from openpyxl import Workbook
from termcolor import colored
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

now = time.strftime("%H:%M:%S", time.localtime(time.time()))

init(autoreset=True) #resetare automana a culorilor la default

def openFile_XLSX(locatie):
#    print('++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
#    print('+                  Alege fisierul dorit                            +')
#    print('++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n')
    print('\nFolederul curent este:\n', Fore.GREEN + locatie)
    dirs = os.listdir(locatie)

    lista_files = []
    for files in dirs:
        lista_files.append(files)
#    print('\n', lista_files)
#   scot doar fisierele Excel
    lista_files_xlsx = [s for s in lista_files if ('.xlsx' or '.XLSX') in s]
    nr_xls = len(lista_files_xlsx)
    print('\nIn folderul curent sunt ', Fore.GREEN + str(nr_xls), ' fisiere Excel...')
    print('Alege fisierul dorit!!!\n')

    if len(lista_files_xlsx) == 0:
        print(Fore.RED + 'In folederul ales nu sunt fisiere .XLSX!')
        print(Fore.RED + 'Ruleaza din nou si alege alt folder!')
        sys.exit() 
    else:
        for x in range(0, nr_xls):
            x += 1
            print('[', x ,'] ', lista_files_xlsx[x-1])

        print('\nAlege indexul fisierul [de la 1 la',nr_xls,']: ')
        print('Trebuie sa fie un fisier cu extensia . XLSX !')
        inp = input()
        inp2 = int(inp)

        while inp2 not in range(1, nr_xls+1):
            print('\nIndexul ',Fore.RED + str(inp2),' nu exista! Ruleaza din nou si alege alt fisier!')
            input('Apasa ENTER pentru iesire!')
            sys.exit()

    print('Fisierul ales este: ', Fore.GREEN + lista_files_xlsx[inp2-1])
    print('\nDaca NU vrei sa pastrezi fisierul, apasa tasta <n> + ENTER ! ')
    inp_str = input('\nPentru continuare apasa ENTER!')

    while inp_str in ['N','n']:
        sys.exit('\nRuleaza din nou si alege alt fisier!')
    
    load = lista_files_xlsx[inp2-1]
    return load
    
def loadExcel(fisier_excel):    
    print(Fore.RED + '\nPutina rabdare, lucrez din greu...')
    wb = openpyxl.load_workbook(fisier_excel)
    print('\nFisierul are urmatoarele file: ')
    sheet_names = wb.get_sheet_names()
    for sheets in sheet_names:
        print(Fore.GREEN + sheets,' ')
    print('-----------------------------------------------------------------')
    return wb

def currentDate():
    data_curenta = datetime.datetime.now().strftime("%d.%m.%Y")
    luna_curenta = datetime.datetime.now().month
    anul_curent = datetime.datetime.now().year
    print('Data curenta: ', Fore.LIGHTYELLOW_EX + data_curenta)
    print('-----------------------------------------------------------------')
    print('Calendarul lunii curente: \n')
    print(Fore.LIGHTYELLOW_EX + calendar.month(anul_curent, luna_curenta))
    print('-----------------------------------------------------------------')
    zi_verificare = int(input('Introdu ziua care trebuie verificata: '))
    my_date = datetime.date(anul_curent, luna_curenta, zi_verificare)
    print('\nZiua selectata este: ', Fore.LIGHTYELLOW_EX + str(my_date.strftime("%d.%m.%Y")), '\n')
    return str(my_date.strftime("%d.%m.%Y"))

def currentWeekOfYear(year, month, day):
    return datetime.date(year, month, day).isocalendar()[1]

def searchShift(schimb):
    lista_schimb = []
    for i in range(1, len(lol_transport)):
        if lol_transport[i][2] == schimb:
            lista_schimb.append(lol_transport[i])
    return lista_schimb

def searchTeam(echipa):
    lista_echipa = []
    for i in range(1, len(lol_transport)):
        if lol_transport[i][2] == echipa:
            lista_echipa.append(lol_transport[i])
    return lista_echipa

def searchTeamShift(data_lucru, echipa):
    schimb = ''
    for i in range(0, len(lol_program)):
        if lol_program[i][0] == data_lucru:
            if echipa == 'ECHIPA 1':
                schimb = str(lol_program[i][1])
            elif echipa == 'ECHIPA 2':
                schimb = str(lol_program[i][2])
            elif echipa == 'ECHIPA 3':
                schimb = str(lol_program[i][3])
            elif echipa == 'ECHIPA 4':
                schimb = str(lol_program[i][4])
    return schimb

def pontajEchipe(schimb, lista_echipa):
    for i in range(len(lista_echipa)):
        lista_echipa[i][2] = schimb

def addTeamShiftsToLists(lista_echipa, lista_sch1, lista_sch2, lista_sch3):
    if lista_echipa[0][2] == '1':
       lista_sch1 += lista_echipa
    elif lista_echipa[0][2] == '2':
        lista_sch2 += lista_echipa
    elif lista_echipa[0][2] == '3':
        lista_sch3.extend(lista_echipa)

def listareMatrice(lol):
    for i in range(0, len(lol)):
        print(lol[i])
    input('\nApasa ENTER pentru continuare\n')
#def writeShiftsToXLSX(shift):
    
#corpul principal al scriptului
wb = Workbook()
print(Fore.RED + "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
print('+                                                                  +')
print('+ Script-ul genereaza fisiere cu tabelele de pontaj zilnic pentru  +')
print('+                 fiecare schimb de la SubGrupuri                  +')
print('+             Start script... ', str(now),'                           +')
print(Fore.GREEN + "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n")


print(Fore.LIGHTCYAN_EX + 'Incarca fisierul de transport...')
#load_file = openFile_XLSX('C:/Users/robert.ban/Documents/Work/_coding/Formular_pontaj_zilnic')
load_file = openFile_XLSX('../Formular_pontaj_zilnic')

wb = loadExcel(load_file)
#print(Fore.LIGHTYELLOW_EX + currentDate())
data_lucru = currentDate()
#print(Fore.LIGHTYELLOW_EX + data_lucru)
#print(wb.active)

lol_transport = []
wb.active = 0
#wb_new = Workbook()
max_rows = wb.active.max_row
max_cols = wb.active.max_column
for i in range(0, max_rows):
    lol_transfer =[]
    for j in range(0, 2):
        lol_transfer.append(wb.active.cell(row=i+1, column=j+1).value)
    lol_transport.append(lol_transfer)

#caut saptamana curenta din an

sapt_curenta_upper = 'WK' + str(currentWeekOfYear(int(data_lucru[6:]), int(data_lucru[3:5]), int(data_lucru[:2])))
sapt_curenta_low = 'wk' + str(currentWeekOfYear(int(data_lucru[6:]), int(data_lucru[3:5]), int(data_lucru[:2])))
index_coloana = 0
for i in range(0, max_rows):
    if (sapt_curenta_upper == str(wb.active.cell(row=1, column=i+1).value)) or (sapt_curenta_low == str(wb.active.cell(row=1, column=i+1).value)):
        print('Pentru ', Fore.LIGHTYELLOW_EX + data_lucru, 'se aplica transportul pt saptamana ', Fore.LIGHTYELLOW_EX + str(wb.active.cell(row=1, column=i+1).value))
        index_coloana = i + 1

for i in range(0, max_rows):
    lol_transport[i].append(str(wb.active.cell(row=i+1, column=index_coloana).value))

#input('baga enter...')

#listez sub forma de matrice
#for i in range(0, len(lol_transport)):
#    print(lol_transport[i], '\n')

#print("\nNumarul de persoane in fisierul de transport: ", Fore.GREEN + str(len(lol_transport)-1))

lista_sch1 = searchShift('1')
lista_sch2 = searchShift('2')
lista_sch3 = searchShift('3')

lista_ech1 = searchTeam('ECHIPA 1')
lista_ech2 = searchTeam('ECHIPA 2')
lista_ech3 = searchTeam('ECHIPA 3')
lista_ech4 = searchTeam('ECHIPA 4')

#Listez situatia pe schimbut=ri si echipe
print('\nsch 1: ', len(lista_sch1), '\nsch 2: ', len(lista_sch2), '\nsch 3: ', len(lista_sch3))
print('--------------------------')
print('Total: ', len(lista_sch1) + len(lista_sch2) + len(lista_sch3))
print('\nech 1: ', len(lista_ech1), '\nech 2: ', len(lista_ech2), '\nech 3: ', len(lista_ech3), '\nech 4: ', len(lista_ech4))
print('--------------------------')
print('Total: ', len(lista_ech1) + len(lista_ech2) + len(lista_ech3) + len(lista_ech4))

#scot programarea echipelor si o salvezi intr-o lista

lol_program = []
wb.active = 1
max_rows = wb.active.max_row
max_cols = wb.active.max_column
for i in range(0, max_rows):
    lol_transfer =[]
    try:
        lol_transfer.append((wb.active.cell(row=i+1, column=1).value).strftime("%d.%m.%Y"))
    except:
        lol_transfer.append(wb.active.cell(row=i+1, column=1).value)
        pass
    for j in range(1, 5):
        lol_transfer.append(wb.active.cell(row=i+1, column=j+1).value)
    lol_program.append(lol_transfer)

#caut schimbul celor care lucreaza in tura continua
schimb = searchTeamShift(data_lucru, 'ECHIPA 1')
pontajEchipe(schimb, lista_ech1)
schimb = searchTeamShift(data_lucru, 'ECHIPA 2')
pontajEchipe(schimb, lista_ech2)
schimb = searchTeamShift(data_lucru, 'ECHIPA 3')
pontajEchipe(schimb, lista_ech3)
schimb = searchTeamShift(data_lucru, 'ECHIPA 4')
pontajEchipe(schimb, lista_ech4)

#listareMatrice(lista_ech1)
#listareMatrice(lista_ech2)
#listareMatrice(lista_ech3)
#listareMatrice(lista_ech4)

#listareMatrice(lista_sch1)

addTeamShiftsToLists(lista_ech1, lista_sch1, lista_sch2, lista_sch3)
addTeamShiftsToLists(lista_ech2, lista_sch1, lista_sch2, lista_sch3)
addTeamShiftsToLists(lista_ech3, lista_sch1, lista_sch2, lista_sch3)
addTeamShiftsToLists(lista_ech4, lista_sch1, lista_sch2, lista_sch3)

#listareMatrice(lista_sch1)

#Listez situatia pe schimburi si echipe
print('\nsch 1: ', len(lista_sch1), '\nsch 2: ', len(lista_sch2), '\nsch 3: ', len(lista_sch3))
print('--------------------------')
print('Total: ', len(lista_sch1) + len(lista_sch2) + len(lista_sch3))

input("baga enter inainte de scrierea datelor in fisier")

#incarc formularul pentru pontaj
print(Fore.LIGHTCYAN_EX + '\nIncarca formularul pentru pontaj...')
wb_formular = Workbook()
#load_formular = openFile_XLSX('../Formular_pontaj_zilnic')
wb_formular = loadExcel('FORMULAR_PONTAJ_SG.xlsx')
#print('Sheet activ formular', wb_formular.active)

#Sch 1
wb_formular.active = 0
wb_formular.active.cell(row=1, column=3).value = data_lucru
wb_formular.active.cell(row=2, column=3).value = 'SCH 1'
for i in range(0, len(lista_sch1)):
    wb_formular.active.cell(row=i+6, column=1).value = i+1
    wb_formular.active.cell(row=i+6, column=2).value = lista_sch1[i][1]

#Sch 2
wb_formular.active = 1
wb_formular.active.cell(row=1, column=3).value = data_lucru
wb_formular.active.cell(row=2, column=3).value = 'SCH 2'
for i in range(0, len(lista_sch2)):
    wb_formular.active.cell(row=i+6, column=1).value = i+1
    wb_formular.active.cell(row=i+6, column=2).value = lista_sch2[i][1]

#Sch 3
wb_formular.active = 2
wb_formular.active.cell(row=1, column=3).value = data_lucru
wb_formular.active.cell(row=2, column=3).value = 'SCH 3'
for i in range(0, len(lista_sch3)):
    wb_formular.active.cell(row=i+6, column=1).value = i+1
    wb_formular.active.cell(row=i+6, column=2).value = lista_sch3[i][1]

#salvez fisierul pe disc
wb_formular.active = 0
file_name_pontaj = 'FORMULAR_PONTAJ_SG_' + data_lucru + '.XLSX'
wb_formular.save(file_name_pontaj)
print('\nAm salvat fisierul de pontaj: ', Fore.GREEN + file_name_pontaj , '\n')


input('Apasa ENTER pentru terminare!')